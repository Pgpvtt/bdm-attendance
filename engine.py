# -*- coding: utf-8 -*-
"""
BDM Attendance Engine
Converts a WhatsApp group chat export (exported WITH media, so photo captions = store names
are preserved) into the standard "APPLE BDM VISIT" Excel workbook + dashboard, and returns
metrics for on-screen display.

Pure logic, no file I/O at import time. Used by app.py (Streamlit) and can be run standalone.
"""
import re, datetime, zipfile, io
from collections import defaultdict

# ---------------- configuration ----------------
# sender name (as it appears in WhatsApp) -> BDM sheet name
DEFAULT_MAP = {
    "Rohith": "ROHITH",
    "Munindra": "MUNINDRA",
    "Mahendra Dhoni": "MAHENDRA",
    "anil": "ANIL",
    "+91 72042 06046": "VINOD",
    "Darshan": "DARSHAN",
}
DEFAULT_ORDER = ["ROHITH", "MUNINDRA", "MAHENDRA", "ANIL", "VINOD", "DARSHAN"]

LEFT_RE = re.compile(r"\bleft\b", re.I)
OFFICE_RE = re.compile(r"\boffice\b", re.I)
_LINE = re.compile(r"^(\d{2})/(\d{2})/(\d{2}), (\d{2}):(\d{2}) - (.*)$")


# ---------------- parsing ----------------
def read_export(file_bytes_or_text):
    """Accepts raw text (str), .txt bytes, or a .zip (bytes) of a WhatsApp export.
    Returns the chat text."""
    if isinstance(file_bytes_or_text, str):
        return file_bytes_or_text
    data = file_bytes_or_text
    # zip?
    if data[:2] == b"PK":
        zf = zipfile.ZipFile(io.BytesIO(data))
        # pick the largest .txt
        txts = [n for n in zf.namelist() if n.lower().endswith(".txt")]
        if not txts:
            raise ValueError("No .txt found inside the zip.")
        txts.sort(key=lambda n: zf.getinfo(n).file_size, reverse=True)
        return zf.read(txts[0]).decode("utf-8", errors="replace")
    return data.decode("utf-8", errors="replace")


def parse(text):
    """-> list of [date 'YYYY-MM-DD', time 'HH:MM', sender, body] (caption continuation joined)."""
    msgs = []
    for ln in text.splitlines():
        m = _LINE.match(ln)
        if m:
            dd, mm, yy, HH, MM, rest = m.groups()
            if ": " in rest:
                sender, body = rest.split(": ", 1)
            else:
                sender, body = None, rest
            msgs.append([f"20{yy}-{mm}-{dd}", f"{HH}:{MM}", sender, body])
        else:
            if msgs:
                msgs[-1][3] += "\n" + ln
    return msgs


def _cap(body):
    return body[len("<Media omitted>"):].strip() if body.startswith("<Media omitted>") else None

def _is_left(c):   return bool(LEFT_RE.search(c))
def _is_office(c): return bool(OFFICE_RE.search(c)) and not _is_left(c)
def _store(c):     return re.sub(r"\s+", " ", c).strip().upper()
def _tparse(t):    h, m = t.split(":"); return datetime.time(int(h), int(m))
def _tmin(t):      h, m = t.split(":"); return int(h) * 60 + int(m)


def build_events(msgs, mapping):
    """Group photo captions per BDM per date, deduping exact repeats (e.g. double-sent photos)."""
    ev = defaultdict(lambda: defaultdict(list))
    last = {}
    unmapped = defaultdict(int)
    for d, t, s, b in msgs:
        if s is None:
            continue
        c = _cap(b)
        if c is None:
            continue
        if s not in mapping:
            unmapped[s] += 1
            continue
        bdm = mapping[s]
        if last.get((s, d)) == (t, c):
            continue
        last[(s, d)] = (t, c)
        ev[bdm][d].append((t, c))
    # sort + dedupe per bdm/day
    for bdm in ev:
        for d in ev[bdm]:
            seen, out = set(), []
            for t, c in sorted(ev[bdm][d]):
                if (t, c) in seen:
                    continue
                seen.add((t, c)); out.append((t, c))
            ev[bdm][d] = out
    # return as plain (picklable) dicts
    return {b: dict(days) for b, days in ev.items()}, dict(unmapped)


def parse_events(file_bytes_or_text, mapping=None):
    """Parse an export once -> (events, unmapped_senders, all_dates). Cheap; cache this."""
    mapping = mapping or DEFAULT_MAP
    text = read_export(file_bytes_or_text)
    ev, unmapped = build_events(parse(text), mapping)
    all_dates = sorted({d for bdm in ev for d in ev[bdm]})
    return ev, unmapped, all_dates


def filter_events(ev, start=None, end=None):
    """Keep only dates within [start, end] (inclusive). start/end = 'YYYY-MM-DD' or None."""
    out = {}
    for bdm, days in ev.items():
        kept = {d: lst for d, lst in days.items()
                if (start is None or d >= start) and (end is None or d <= end)}
        out[bdm] = kept
    return out


def reconstruct(day):
    """Day events -> ordered visits: dict(store, tin, tout, office)."""
    visits, cur = [], None
    for t, c in sorted(day):
        if not c:
            continue
        if _is_office(c):
            if cur:
                visits.append({"store": cur[0], "tin": cur[1], "tout": None, "office": False})
            visits.append({"store": "OFFICE", "tin": t, "tout": None, "office": True})
            cur = None
            continue
        if _is_left(c):
            if cur:
                visits.append({"store": cur[0], "tin": cur[1], "tout": t, "office": False})
                cur = None
        else:
            st = _store(c)
            if cur:
                if cur[0] == st:
                    continue
                visits.append({"store": cur[0], "tin": cur[1], "tout": None, "office": False})
            cur = (st, t)
    if cur:
        visits.append({"store": cur[0], "tin": cur[1], "tout": None, "office": False})
    return visits


def store_count(visits):
    return sum(1 for v in visits if not v["office"])


# ---------------- metrics ----------------
def hm(mins):
    if mins is None:
        return "-"
    return f"{int(mins) // 60}:{int(mins) % 60:02d}"


def store_coverage(ev, order):
    """Aggregate visits per store across the (filtered) data."""
    agg = {}
    for b in order:
        for d in sorted(ev.get(b, {})):
            for v in reconstruct(ev[b][d]):
                if v["office"]:
                    continue
                a = agg.setdefault(v["store"], {"visits": 0, "bdms": set(), "first": d, "last": d})
                a["visits"] += 1; a["bdms"].add(b)
                a["first"] = min(a["first"], d); a["last"] = max(a["last"], d)
    rows = [{"store": st, "visits": a["visits"], "bdms": ", ".join(sorted(a["bdms"])),
             "nbdm": len(a["bdms"]), "first": a["first"], "last": a["last"]} for st, a in agg.items()]
    rows.sort(key=lambda r: (-r["visits"], r["store"]))
    return rows


def anomalies(ev, order):
    """Return (issue_rows, quality_rows).
    issue_rows: per-visit problems (duplicate / too-short / too-long).
    quality_rows: per-BDM missing-exit ('left' photo) rate."""
    issues = []
    quality = []
    for b in order:
        tot = miss = 0
        for d in sorted(ev.get(b, {})):
            visits = [v for v in reconstruct(ev[b][d]) if not v["office"]]
            seen = set()
            for v in visits:
                tot += 1
                if v["tout"] is None:
                    miss += 1
                else:
                    dur = _tmin(v["tout"]) - _tmin(v["tin"])
                    if dur < 3:
                        issues.append((d, b, v["store"], f"Very short visit ({dur} min) — possible mis-tag"))
                    elif dur > 180:
                        issues.append((d, b, v["store"], f"Very long visit ({hm(dur)}) — maybe forgot 'left'"))
                if v["store"] in seen:
                    issues.append((d, b, v["store"], "Same store logged twice that day"))
                seen.add(v["store"])
        pct = round(100 * miss / tot) if tot else 0
        quality.append({"bdm": b, "visits": tot, "missing_exit": miss, "pct": pct})
    issues.sort()
    return issues, quality


def compute(ev, order):
    all_dates = sorted({d for bdm in ev for d in ev[bdm]})
    latest = all_dates[-1] if all_dates else None
    per = {}
    for bdm in order:
        days_worked = leave = total = field = 0
        logins = []
        per_day = {}
        for d in sorted(ev.get(bdm, {})):
            v = reconstruct(ev[bdm][d]); sc = store_count(v)
            per_day[d] = sc
            reals = [x for x in v if not x["office"]]
            if reals:
                days_worked += 1; total += sc
                logins.append(min(_tmin(x["tin"]) for x in reals))
                field += sum(max(0, _tmin(x["tout"]) - _tmin(x["tin"])) for x in reals if x["tout"])
            else:
                leave += 1
        per[bdm] = dict(days_worked=days_worked, leave=leave, total=total, field=field,
                        avg_login=(sum(logins)//len(logins) if logins else None),
                        avg_stores=(round(total/days_worked, 1) if days_worked else 0),
                        avg_store_min=(round(field/total) if total else 0),
                        per_day=per_day)
    return all_dates, latest, per


# ---------------- Excel build ----------------
def build_workbook(ev, order):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HDR = PatternFill("solid", fgColor="CFE2F3"); TOT = PatternFill("solid", fgColor="C9DAF8")
    CNT = PatternFill("solid", fgColor="FFF2CC"); NAVY = "1F3864"; NFILL = PatternFill("solid", fgColor=NAVY)
    thin = Side("thin", color="000000"); med = Side("medium", color="000000")
    def bd(l=thin, r=thin, t=thin, b=thin): return Border(left=l, right=r, top=t, bottom=b)
    F10 = Font(name="Calibri", size=10); F10B = Font(name="Calibri", size=10, bold=True)
    F11 = Font(name="Calibri", size=11); WHB = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    CEN = Alignment(horizontal="center"); LEFT = Alignment(horizontal="left")
    HEAD = ["Sl. No", "DATE / DAY", "STORE", "TIME IN", "TIME OUT", "STORE TIME"]

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    all_dates, latest, per = compute(ev, order)

    # ---- dashboard ----
    ds = wb.create_sheet("DASHBOARD"); ds.sheet_view.showGridLines = False
    ds.cell(row=1, column=1, value="ACPL — BDM FIELD VISIT DASHBOARD").font = Font(size=16, bold=True, color=NAVY)
    dtl = datetime.datetime.strptime(latest, "%Y-%m-%d") if latest else None
    ds.cell(row=2, column=1, value=f"Auto-generated from WhatsApp group  |  Data through "
            f"{dtl.strftime('%d-%b-%Y (%A)') if dtl else '-'}").font = Font(size=10, color="595959")
    def dhead(row, cols, c0=1):
        for j, h in enumerate(cols):
            c = ds.cell(row=row, column=c0+j, value=h); c.fill = NFILL; c.font = WHB; c.alignment = CEN; c.border = bd()
    # today
    ds.cell(row=4, column=1, value=f"TODAY  —  {dtl.strftime('%d-%b-%Y') if dtl else '-'}").font = Font(size=12, bold=True, color=NAVY)
    dhead(5, ["BDM", "Status", "Stores Today", "First In", "Last Out", "Field Time"])
    r = 6; tot_today = 0
    for b in order:
        v = reconstruct(ev.get(b, {}).get(latest, [])); reals = [x for x in v if not x["office"]]
        ds.cell(row=r, column=1, value=b).font = F10B; ds.cell(row=r, column=1).border = bd()
        if reals:
            ins = [_tmin(x["tin"]) for x in reals]; outs = [_tmin(x["tout"]) for x in reals if x["tout"]]
            fm = sum(max(0, _tmin(x["tout"]) - _tmin(x["tin"])) for x in reals if x["tout"])
            vals = ["Working", len(reals), hm(min(ins)), hm(max(outs)) if outs else "-", hm(fm)]; tot_today += len(reals)
        else:
            vals = ["On Leave", 0, "-", "-", "-"]
        for j, val in enumerate(vals, 2):
            cc = ds.cell(row=r, column=j, value=val); cc.alignment = CEN; cc.border = bd()
            if j == 2: cc.fill = PatternFill("solid", fgColor="C6EFCE" if val == "Working" else "FFC7CE")
        r += 1
    ds.cell(row=r, column=1, value="TEAM TOTAL").font = F10B
    for c in range(1, 7): ds.cell(row=r, column=c).fill = TOT; ds.cell(row=r, column=c).border = bd()
    ds.cell(row=r, column=3, value=tot_today).font = F10B; ds.cell(row=r, column=3).alignment = CEN
    # period
    pr = r + 3
    ds.cell(row=pr-1, column=1, value="PERIOD SUMMARY").font = Font(size=12, bold=True, color=NAVY)
    dhead(pr, ["BDM", "Days Worked", "Days Leave", "Total Stores", "Avg Stores/Day", "Total Field Time", "Avg Time/Store", "Avg First-In"])
    r = pr + 1; T = defaultdict(int)
    for b in order:
        m = per[b]
        vals = [b, m["days_worked"], m["leave"], m["total"], m["avg_stores"], hm(m["field"]), hm(m["avg_store_min"]), hm(m["avg_login"])]
        for j, val in enumerate(vals, 1):
            cc = ds.cell(row=r, column=j, value=val); cc.alignment = (LEFT if j == 1 else CEN); cc.border = bd()
            if j == 1: cc.font = F10B
        T["dw"] += m["days_worked"]; T["lv"] += m["leave"]; T["ts"] += m["total"]; T["fm"] += m["field"]; r += 1
    for j, val in enumerate(["TEAM", T["dw"], T["lv"], T["ts"], "", hm(T["fm"]), "", ""], 1):
        cc = ds.cell(row=r, column=j, value=val); cc.font = F10B; cc.fill = TOT; cc.alignment = (LEFT if j == 1 else CEN); cc.border = bd()
    # last 14 days matrix
    mr = r + 3
    ds.cell(row=mr-1, column=1, value="DAILY STORE VISITS (last 14 days)").font = Font(size=12, bold=True, color=NAVY)
    dhead(mr, ["Date"] + order + ["Team"]); rr = mr + 1
    for d in all_dates[-14:]:
        dd = datetime.datetime.strptime(d, "%Y-%m-%d")
        ds.cell(row=rr, column=1, value=dd.strftime("%d-%b %a")).font = F10B; ds.cell(row=rr, column=1).border = bd()
        tt = 0
        for j, b in enumerate(order, 2):
            sc = per[b]["per_day"].get(d)
            cc = ds.cell(row=rr, column=j, value=(sc if sc is not None else "")); cc.alignment = CEN; cc.border = bd()
            if not sc: cc.fill = PatternFill("solid", fgColor="F2F2F2")
            if isinstance(sc, int): tt += sc
        tc = ds.cell(row=rr, column=2+len(order), value=tt); tc.font = F10B; tc.alignment = CEN; tc.fill = TOT; tc.border = bd(); rr += 1
    for i, w in enumerate([16] + [12]*len(order) + [10]):
        ds.column_dimensions[openpyxl.utils.get_column_letter(1+i)].width = w

    # ---- STORE COVERAGE sheet ----
    cov = store_coverage(ev, order)
    cs = wb.create_sheet("STORE COVERAGE"); cs.sheet_view.showGridLines = False
    cs.cell(row=1, column=1, value="STORE COVERAGE (selected period)").font = Font(size=13, bold=True, color=NAVY)
    cs.cell(row=2, column=1, value=f"{len(cov)} unique stores visited  ·  "
            f"{sum(c['visits'] for c in cov)} total visits").font = Font(size=10, color="595959")
    ch = ["#", "Store", "Visits", "Visited by", "First Visit", "Last Visit"]
    for j, h in enumerate(ch, 1):
        c = cs.cell(row=4, column=j, value=h); c.fill = NFILL; c.font = WHB; c.alignment = CEN; c.border = bd()
    for i, c in enumerate(cov, 1):
        rr = 4 + i
        for j, val in enumerate([i, c["store"], c["visits"], c["bdms"],
                                 datetime.datetime.strptime(c["first"], "%Y-%m-%d"),
                                 datetime.datetime.strptime(c["last"], "%Y-%m-%d")], 1):
            cell = cs.cell(row=rr, column=j, value=val); cell.border = bd()
            cell.font = F10B if j == 2 else F10
            if j in (5, 6): cell.number_format = "dd-mmm"; cell.alignment = CEN
            if j in (1, 3): cell.alignment = CEN
            if j == 3 and c["visits"] == 1: cell.fill = PatternFill("solid", fgColor="FFF2CC")  # visited once
    for col, w in zip("ABCDEF", [5, 40, 8, 26, 11, 11]): cs.column_dimensions[col].width = w
    cs.freeze_panes = "A5"; cs.auto_filter.ref = f"A4:F{4+len(cov)}"

    # ---- DATA QUALITY sheet ----
    issues, quality = anomalies(ev, order)
    qs = wb.create_sheet("DATA QUALITY"); qs.sheet_view.showGridLines = False
    qs.cell(row=1, column=1, value="DATA QUALITY & ANOMALIES").font = Font(size=13, bold=True, color=NAVY)
    qs.cell(row=3, column=1, value="Per-BDM: how often the 'left' (exit) photo is missing").font = Font(size=11, bold=True, color=NAVY)
    for j, h in enumerate(["BDM", "Visits", "Missing 'left'", "Missing %"], 1):
        c = qs.cell(row=4, column=j, value=h); c.fill = NFILL; c.font = WHB; c.alignment = CEN; c.border = bd()
    for i, q in enumerate(quality, 1):
        rr = 4 + i
        for j, val in enumerate([q["bdm"], q["visits"], q["missing_exit"], f"{q['pct']}%"], 1):
            cell = qs.cell(row=rr, column=j, value=val); cell.border = bd(); cell.alignment = (LEFT if j == 1 else CEN)
            if j == 1: cell.font = F10B
            if j == 4: cell.fill = PatternFill("solid", fgColor="FFC7CE" if q["pct"] >= 70 else ("FFF2CC" if q["pct"] >= 30 else "C6EFCE"))
    base = 4 + len(quality) + 2
    qs.cell(row=base, column=1, value=f"Specific anomalies to review ({len(issues)})").font = Font(size=11, bold=True, color=NAVY)
    for j, h in enumerate(["Date", "BDM", "Store", "Issue"], 1):
        c = qs.cell(row=base+1, column=j, value=h); c.fill = NFILL; c.font = WHB; c.alignment = CEN; c.border = bd()
    for i, (d, b, store, issue) in enumerate(issues, 1):
        rr = base + 1 + i
        for j, val in enumerate([datetime.datetime.strptime(d, "%Y-%m-%d"), b, store, issue], 1):
            cell = qs.cell(row=rr, column=j, value=val); cell.border = bd()
            if j == 1: cell.number_format = "dd-mmm-yy"
            if j == 2: cell.font = F10B
    for col, w in zip("ABCD", [12, 12, 38, 44]): qs.column_dimensions[col].width = w

    # ---- per-BDM sheets ----
    def write_header(ws, row):
        for c, txt in enumerate(HEAD, 1):
            cell = ws.cell(row=row, column=c, value=txt); cell.fill = HDR; cell.font = F10B
            cell.alignment = CEN if c >= 4 else LEFT
            cell.border = bd(med if c == 1 else thin, med if c == 6 else thin, med, thin)
    for b in order:
        ws = wb.create_sheet(b)
        for col, w in zip("ABCDEF", [7.3, 12, 40, 10, 12, 12]):
            ws.column_dimensions[col].width = w
        r = 1
        for d in sorted(ev.get(b, {})):
            visits = reconstruct(ev[b][d]); dt = datetime.datetime.strptime(d, "%Y-%m-%d"); wk = dt.strftime("%A").upper()
            write_header(ws, r); r += 1
            if not visits:
                for i, lab in enumerate([dt, wk], 1):
                    ws.cell(row=r, column=1, value=i).font = F10; ws.cell(row=r, column=1).border = bd(med)
                    bc = ws.cell(row=r, column=2, value=lab); bc.font = F10B; bc.alignment = LEFT; bc.border = bd()
                    if i == 1: bc.number_format = "mm-dd-yy"; ws.cell(row=r, column=3, value="-").font = F11
                    for c in range(3, 7): ws.cell(row=r, column=c).border = bd()
                    r += 1
                total = "ON LEAVE "
            else:
                for i, v in enumerate(visits, 1):
                    ws.cell(row=r, column=1, value=i).font = F10; ws.cell(row=r, column=1).border = bd(med)
                    bc = ws.cell(row=r, column=2, value=(dt if i == 1 else (wk if i == 2 else None)))
                    bc.font = F10B; bc.alignment = LEFT; bc.border = bd()
                    if i == 1: bc.number_format = "mm-dd-yy"
                    ws.cell(row=r, column=3, value=v["store"]).font = F11; ws.cell(row=r, column=3).border = bd()
                    din = ws.cell(row=r, column=4, value=_tparse(v["tin"])); din.font = F11; din.number_format = "h:mm"; din.alignment = CEN; din.border = bd()
                    if v["tout"]:
                        de = ws.cell(row=r, column=5, value=_tparse(v["tout"])); de.number_format = "h:mm"; fc = ws.cell(row=r, column=6, value=f"=E{r}-D{r}")
                    else:
                        de = ws.cell(row=r, column=5, value="NO UPDATE"); fc = ws.cell(row=r, column=6, value="-")
                    de.font = F11; de.alignment = CEN; de.border = bd(); fc.font = F10; fc.number_format = "h:mm;@"; fc.alignment = CEN; fc.border = bd()
                    r += 1
                total = f"{store_count(visits)} STORES"
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            tc = ws.cell(row=r, column=1, value="TOTAL"); tc.fill = TOT; tc.font = F10B; tc.border = bd()
            ws.cell(row=r, column=2).fill = TOT; ws.cell(row=r, column=2).border = bd()
            cc = ws.cell(row=r, column=3, value=total); cc.fill = CNT; cc.font = F10B; cc.border = bd()
            for c in (4, 5, 6): ws.cell(row=r, column=c).border = bd()
            r += 1
    return wb, (all_dates, latest, per)


def process(file_bytes_or_text, mapping=None, order=None):
    """High-level: export -> (Workbook, events, metrics, unmapped_senders)."""
    mapping = mapping or DEFAULT_MAP
    order = order or DEFAULT_ORDER
    text = read_export(file_bytes_or_text)
    msgs = parse(text)
    ev, unmapped = build_events(msgs, mapping)
    wb, metrics = build_workbook(ev, order)
    return wb, ev, metrics, unmapped


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else \
        r"C:\Users\AHCPL\Desktop\afreen work\WhatsApp Chat with ACPL BDM's ATTENDANCE.zip"
    with open(path, "rb") as f:
        wb, ev, (all_dates, latest, per), unmapped = process(f.read())
    out = r"C:\Users\AHCPL\Desktop\afreen work\APPLE BDM VISIT - GENERATED.xlsx"
    wb.save(out)
    print("Saved", out)
    print("Dates:", all_dates[0], "->", all_dates[-1], "| BDMs:", list(per.keys()))
    print("Unmapped senders (ignored):", {k: v for k, v in sorted(unmapped.items(), key=lambda x: -x[1])})
